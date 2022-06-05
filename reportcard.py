'''
import openpyxl
w1=openpyxl.load_workbook('C:\\Users\\Tanansh Ahuja\\Desktop\\report2.xlsx')


for i in range(1,45):
    sheet = w1[str(i)]
    sheet.cell(15,15).value=''
    sheet.cell(15,16).value=''
    sheet.cell(15,17).value=''
    sheet.cell(15,18).value=''
    sheet.cell(15,19).value=''
    sheet.cell(15,20).value=''

w1.save('C:\\Users\\Tanansh Ahuja\\Desktop\\report2.xlsx')

'''
'''
import win32com.client
from pywintypes import com_error


# Path to original excel file
WB_PATH = 'C:\\hoge\\fuga\\report2.xlsx'
# PDF path when saving
PATH_TO_PDF = 'C:\\hoge\\fuga\\report2.pdf'


excel = win32com.client.Dispatch("Excel.Application")

excel.Visible = False

try:
    print('Start conversion to PDF')

    # Open
    wb = excel.Workbooks.Open(WB_PATH)

    # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
    ws_index_list = list(range(1,45))
    wb.WorkSheets(ws_index_list).Select()

    # Save
    wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
except com_error as e:
    print('failed.')
else:
    print('Succeeded.')
finally:
    wb.Close()
    excel.Quit()
'''
import win32com.client
o = win32com.client.Dispatch("Excel.Application")
o.Visible = False
wb_path = r'c:\user\desktop\report2.xlsx'
wb = o.Workbooks.Open(wb_path)
ws_index_list = [1,4,5] #say you want to print these sheets
path_to_pdf = r'C:\user\desktop\sample.pdf'
print_area = 'A1:M49'
for index in ws_index_list:
    #off-by-one so the user can start numbering the worksheets at 1
    ws = wb.Worksheets[index - 1]
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesTall = 1
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.PrintArea = print_area

wb.WorkSheets(ws_index_list).Select()
wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
