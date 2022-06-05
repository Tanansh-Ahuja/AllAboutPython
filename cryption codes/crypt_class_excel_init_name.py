from random import *
import openpyxl
import datetime


 #this class will have the function for crypting
class __crypt__:                                                       

    def __init__(self,argcount,arga):
        self.count=argcount
        self.a=arga

#this function will create a dictionary with every character linking to one other character   
    def dict_create(self):                                                 
        seed(self.count)                                                    #pattern in randomness
        shuffle(self.a)
        key=self.a.copy()
        shuffle(self.a)
        value=self.a.copy()
        final=dict(zip(key,value))                                      #a dictionary is created
        self.count=self.count+1                                         #count changes for further coading
        return final                                                                #returning to main() output variable

    #crypting time
    def crypt(self,user,b):                                                            #using the dictionary it will crypt what user entered
        output=""
        for i in user:
            output=output+b.get(i)
        return output

#calling everything
class main:                                             #a main class that contains everything

    def __init__(self1,argcount,arga):
        self1.count=argcount
        self1.a=arga
        
    def main(self1,sht,n):
        obj1=__crypt__  (self1.count,self1.a)                                                #an object of class crypt to create to call functions
#Saving the data in an excel file
        sht.cell(n,1).value="DATE"
        sht.cell(n,2).value="TIME"
        sht.cell(n,3).value="ORIGINAL"
        sht.cell(n,4).value="CRYPTED"

        while 1:                                                                
            b=obj1.dict_create()                                    #callind dict_create to make a dictionary
            user=input("Enter: ")                                   
            sht.cell(n+1,1).value=datetime.datetime.now().strftime("%d-%m-%Y")   #save the date
            sht.cell(n+1,2).value=datetime.datetime.now().strftime("%H:%M:%S")  #save the time
            sht.cell(n+1,3).value=user  #saves what the user originally entered
            output=obj1.crypt(user,b)                              #calling crpyt()
            sht.cell(n+1,4).value=output
            wb.save("cry.xlsx")     #saves the file everytime it writes on it
            n=n+1
            print(output)

#main
if __name__=="__main__":
    count =1
    a=list(map(chr,map(int,range(32,127))))
    obj2=main(count,a)
    #loading an excel workbook
    wb = openpyxl.load_workbook("C:\\Users\\com\\Desktop\\PYTHON\\cryption codes\\cry.xlsx")
    #creating a sheet object
    sheet = wb.active
    if(sheet.max_row==1):
        n=sheet.max_row
    else:
        n=sheet.max_row+2  #this was done to give a row space when user re-enters data
    obj2.main(sheet,n)


