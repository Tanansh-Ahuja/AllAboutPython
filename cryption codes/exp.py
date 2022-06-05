from random import *
import openpyxl
import datetime
from tkinter import *


 #this class will have the function for crypting
class __crypt__:                                                       

    def __init__(self,argcount,arga):
        self.count=argcount
        self.a=arga

#this function will create a dictionary with every character linking to one other character   
    def dict_create(self):                                                 
        seed(self.count)                                                    #pattern in randomness
        shuffle(self.a)
        key=self.a.copy()
        shuffle(self.a)
        value=self.a.copy()
        final=dict(zip(key,value))                                      #a dictionary is created
        self.count=self.count+1                                         #count changes for further coading
        return final                                                                #returning to main() output variable

    #crypting time
    def crypt(self,user,b):                                                            #using the dictionary it will crypt what user entered
        output=""
        for i in user:
            output=output+b.get(i)
        return output

#calling everything
class main:                                             #a main class that contains everything

    def __init__(self1,argcount,arga):
        self1.count=argcount
        self1.a=arga
        
    def main(self1,sht,n,root):
        obj1=__crypt__  (self1.count,self1.a)                                                #an object of class crypt to create to call functions
#Saving the data in an excel file
        sht.cell(n,1).value="DATE"
        sht.cell(n,2).value="TIME"
        sht.cell(n,3).value="ORIGINAL"
        sht.cell(n,4).value="CRYPTED"

        while 1:                                                                
            b=obj1.dict_create()                                    #callind dict_create to make a dictionary
            user=input("Enter: ")                                   
            sht.cell(n+1,1).value=datetime.datetime.now().strftime("%d-%m-%Y")   #save the date
            sht.cell(n+1,2).value=datetime.datetime.now().strftime("%H:%M:%S")  #save the time
            sht.cell(n+1,3).value=user  #saves what the user originally entered
            output=obj1.crypt(user,b)                              #calling crpyt()
            sht.cell(n+1,4).value=output
            wb.save("cry.xlsx")     #saves the file everytime it writes on it
            n=n+1
            print(output)
            #root.update()


class __tk__:

    def create():
         #try:
        root=Tk()
        print('start')
        root.geometry('700x700')
        #root.mainloop()
        root.title('Experiments')
        f1=Frame(root,padx=5,pady=5,bg='red',borderwidth=8,relief=SUNKEN,height=300)
        f1.pack(side=TOP,fill=X,padx=5,pady=5)
        f2=Frame(root,padx=5,pady=5,bg='orange',borderwidth=8,relief=RIDGE)
        f2.pack(side=TOP,fill=X,padx=5,pady=5)
        l1=Label(f1,padx=2,pady=2,text='Label 1',font=('Times new roman',20,'bold'),borderwidth=2,relief=RAISED)
        l1.pack(side=TOP)
        l2=Label(f2,padx=2,pady=2,text='Label 2',font=('Times new roman',20,'bold'),borderwidth=2,relief=RAISED)
        l2.pack(side=TOP)
        b1=Button(f2,text='press me',height=2)
        b1.pack(side=TOP)
        b1.bind('<Button-1>',func)
        b1.bind('<Button-3>',reset)
        print('packing done')

        st1=StringVar()
        st1.set('')
        st2=StringVar()
        st2.set('')
        e1=Entry(f1,textvariable=st1)
        e1.pack(side=TOP,fill=X)
        e2=Entry(f1,textvariable=st2)
        e2.pack(side=TOP,fill=X)
        root.update()

    def func(event):
        global st1
        global st2
        #global e1
        #global e2
        #st2.set('')
        #st1.set('')
        #e1.update()
        #e2.update()
        print('in func')
        b=''
        a=st1.get()
        print(a)
        for i in a:
            b=b+chr(ord(i)+1)
        e2.insert(0,b)
        e2.update()
        print(b)


    def reset(event):
        global st1
        global st2
        st1.set('')
        st2.set('')
        e1.update()
        e2.update()

#main
if __name__=="__main__":

    t=__tk__()
    t.create()
    count =1
    a=list(map(chr,map(int,range(32,127))))
    obj2=main(count,a)
    #loading an excel workbook
    wb = openpyxl.load_workbook("C:\\Users\\com\\Desktop\\PYTHON\\cryption codes\\cry1.xlsx")
    #creating a sheet object
    sheet = wb.active
    if(sheet.max_row==1):
        n=sheet.max_row
    else:
        n=sheet.max_row+2  #this was done to give a row space when user re-enters data
    obj2.main(sheet,n,root)
    #finally:
        #root.mainloop()


