from tkinter import *
from playsound import playsound
from time import *
from PIL import ImageTk, Image
import openpyxl

class datasaving:

    def __init__(selfds):
        pass

    


    def save_signup(selfds,n,e,m,p,sheet,mx,wb):
        sheet.cell(mx+1,1).value=n
        sheet.cell(mx+1,2).value=e
        sheet.cell(mx+1,3).value=m
        sheet.cell(mx+1,4).value=p
        wb.save('users.xlsx')
        print('Data saved successully!')

    def pay(selfds):
        pass

    def add(selfds):
        pass

    def money_toggle(selfds):
        pass

    def loan(selfds):
        pass

    def loan_repayment(selfds):
        pass
  



class dataretrival:

    def __init__(selfdr):
        pass



class binding:

    def __init__(selfbind):
       pass

    def focusin(selfbind,event):
        event.widget.delete(0,END)
        event.widget.config(fg='black')

    def focusinpaswrd(selfbind,event):
        event.widget.delete(0,END)
        event.widget.config(fg='black',show='*')

    def destroydbox(selfbind,event):
        selfbind.rootx.destroy()

    def login_check(selfbind,event):

        wb=openpyxl.load_workbook('users.xlsx')
        sheet=wb.active
        mx=sheet.max_row
        print('no of rows: ',mx)
        enroll=''

        e=selfbind.login_enroll.get()
        p=selfbind.login_password.get()

        for i in range(2,mx+1,1):
            if sheet.cell(i,2).value==e:
                enroll=sheet.cell(i,2).value
                paswrd=sheet.cell(i,4).value
                selfbind.name=sheet.cell(i,1).value
                break
           
        print(e)
        print(p)

        if ((e != ''and e!='Enrollment No.') and (p != '' and p!='Password')):
            print('stage 1 cleared')

            if e=='administrator' and p=='administrator':
                print('you are in administrator window')
                selfbind.froot_login_signup.destroy()
                selfbind.administrator()

            elif e==enroll and p==paswrd:
                print('you can access your account ',selfbind.name)
                selfbind.froot_login_signup.destroy()
                selfbind.user()

            else:
                selfbind.error('DATA ERROR! \nThe data you entered doesnt match')

        else:
            selfbind.error('ENTER DATA TO PROCEED!')

    def signup_check(selfbind,event):

        wb=openpyxl.load_workbook('users.xlsx')
        sheet=wb.active
        mx=sheet.max_row

        n=selfbind.signup_name.get()
        e=selfbind.signup_enroll.get()
        m=selfbind.signup_mobile.get()
        p=selfbind.signup_password.get()
        cp=selfbind.signup_cnfpassword.get()

        print('name: ',n)
        print('enroll: ',e)
        print('mobile: ',m)
        print('password: ',p)
        print('confirm password: ',cp)

        nlist=[]
        elist=[]
        for i in range(2,mx+1,1):
            nlist.append(sheet.cell(i,1).value)
            elist.append(sheet.cell(i,2).value)

        if((n != 'Name' and n != '') and (e != 'Enrollment no.' and e !='') and (m != 'Mobile number' and m !='')and (p != 'Password.' and p !='')and (cp != 'Confirm Password' and cp !='')):

            if p==cp:

                if(len(m)==10  and m.isnumeric()):

                    if(e.isnumeric()):

                        if ((n in nlist) and (e in elist)):
                            selfbind.error('This details have been entered before,\n please enter some other details')

                        else:
                            selfbind.save_signup(n,e,m,p,sheet,mx,wb)
                    else:
                        selfbind.error('The Enrollment number\nhs to be numeric')
                else:
                    selfbind.error('The mobile number\nhas to be atleast 10 digits long')
            else:
                selfbind.error('The Passwords dont match! \n Re-enter the password')
        else:
            selfbind.error('Please fill all the \ndialogue boxes to signup')
                    
            
            


class transition:

    def __init__(selft):
        pass

    def des_first_page(selft):
        playsound('money manager.mp3')
        selft.frootfirst_page .destroy()
        selft.login_signup()



class creation(transition,binding,dataretrival,datasaving):

    def __init__(selfc,rootarg):
        selfc.root=rootarg
        transition.__init__(selfc)
        binding.__init__(selfc)
        dataretrival.__init__(selfc)
        datasaving.__init__(selfc)


    def error(selfc,txt):
        selfc.rootx=Tk()
        selfc.rootx.title('ERROR!!!')
        selfc.rootx.geometry('300x100+500+300')
        selfc.rootx.maxsize(300,100)
        selfc.rootx.minsize(300,100)
        l1=Label(selfc.rootx,text=txt, bg='white',font=('Calibri',15))
        l1.pack(side=TOP,fill=X)
        b1=Button(selfc.rootx,text="Ok",relief=RAISED,borderwidth=5,bg='light grey',padx=20,pady=2)
        b1.pack(padx=5,pady=10)
        b1.bind('<Button-1>',selfc.destroydbox)
        selfc.rootx.mainloop()
        

    def first_page(selfc):
        
        i=0
        while i<=1:
            root.attributes('-alpha',i)
            root.update()
            sleep(0.01)
            i=i+0.02
        
        selfc.frootfirst_page=Frame(root,bg='black')
        selfc.frootfirst_page.pack(fill=BOTH,expand=True)
        f1=Frame(selfc.frootfirst_page,padx=5,pady=5)
        f1.pack(anchor=NE)
        l1=Label(selfc.frootfirst_page,text='Created By\nTanansh Ahuja',bg='black',fg='yellow',font=('Times New Roman',15,'bold'))
        l1.pack(anchor='ne')
        l2=Label(selfc.frootfirst_page,text='Welcome\nTo\nThe Money Manager',bg='black',fg='#FF3636',font=('Jester',30,'bold'))
        l2.place(relx=0.5,rely=0.5,anchor=CENTER)
        l2.after(500,selfc.des_first_page)
        root.mainloop()

    def login_signup(selfc):

        selfc.froot_login_signup=Frame(root,bg='white')
        selfc.froot_login_signup.pack(fill=BOTH,expand=True)
        # Making all the frames, atleast as much as possible
        f1=Frame(selfc.froot_login_signup,bg='sky blue',padx=10)
        f1.pack(side=TOP,fill=X,padx=2,pady=2)
        f2=Frame(selfc.froot_login_signup,bg='red')
        f2.pack(side=TOP,fill=BOTH,padx=2,pady=2,expand=1)
        f3=Frame(f2,bg='#F5F5F5',padx=5,pady=5)
        f3.pack(side=LEFT,fill=BOTH,expand=True)
        f4=Frame(f2,bg='#F5F5F5',padx=5,pady=5)
        f4.pack(side=LEFT,fill=BOTH,expand=True)
        # Now we will make login frame
        #variables defination
        selfc.login_enroll=StringVar()
        selfc.login_enroll.set('')
        selfc.login_password=StringVar()
        selfc.login_password.set('')
        # Layout of login
        l1=Label(f1,bg='blue',text='Money Manager',fg='white',padx=10,pady=5,relief=RAISED,borderwidth=5,font=('Algerian',20,'underline'))
        l1.pack(side=LEFT)
        f5=Frame(f1,bg='sky blue',padx=2,pady=20)
        f5.pack(anchor=E,fill=BOTH,expand=1)
        b1=Button(f5,text='Log In',fg='white',bg='blue',padx=3,pady=2,font=('Calibri',10,'bold'),relief=RAISED,borderwidth=5)
        b1.pack(side=RIGHT, padx=15)
        b1.bind('<Button-1>',selfc.login_check)
        e2=Entry(f5,textvariable=selfc.login_password,width=20,fg='grey',font=('times new roman',15,'bold'))
        e2.insert(0,'Password')
        e2.pack(side=RIGHT,padx=10)
        e2.bind('<FocusIn>',selfc.focusinpaswrd)
        e1=Entry(f5,textvariable=selfc.login_enroll,width=20,fg='grey',font=('times new roman',15,'bold'))
        e1.insert(0,'Enrollment No.')
        e1.pack(side=RIGHT,padx=10)
        e1.bind('<FocusIn>',selfc.focusin)
        # Now we make signup page
        #Variable definations
        selfc.signup_name=StringVar()
        selfc.signup_enroll=StringVar()
        selfc.signup_mobile=StringVar()
        selfc.signup_password=StringVar()
        selfc.signup_cnfpassword=StringVar()
        selfc.signup_name.set('')
        selfc.signup_enroll.set('')
        selfc.signup_mobile.set('')
        selfc.signup_password.set('')
        selfc.signup_cnfpassword.set('')
        #layout
        l4=Label(f4,text='New user,huh?',font=('Calibri',15,'bold'),bg='#F5F5F5')
        l4.pack(side=TOP,padx=10,pady=20,anchor=W)
        l5=Label(f4,text='Create a new Account',font=('Calibri',30,'bold'),bg='#F5F5F5')
        l5.pack(side=TOP,padx=10,pady=1,anchor=W)
        f6=Frame(f4,bg='#F5F5F5',padx=2,pady=20)
        f6.place(relx=0.5,rely=0.5,anchor=CENTER)
        f61=Frame(f6,bg='#F5F5F5',padx=5,pady=5)
        f61.pack(side=TOP)
        f611=Frame(f61,bg='#F5F5F5')
        f611.pack(side=TOP,padx=20,pady=10)
        e3=Entry(f611,textvariable=selfc.signup_name,width=12,fg='grey',font=('times new roman',25,'bold'))
        e3.pack(side=LEFT,padx=25)
        e3.insert(0,'Name')
        e3.bind('<FocusIn>',selfc.focusin)
        e4=Entry(f611,textvariable=selfc.signup_enroll,width=12,fg='grey',font=('times new roman',25,'bold'))
        e4.pack(side=LEFT,padx=25)
        e4.insert(0,'Enrollment no.')
        e4.bind('<FocusIn>',selfc.focusin)
        e5=Entry(f61,textvariable=selfc.signup_mobile,width=15,fg='grey',font=('times new roman',25,'bold'))
        e5.pack(side=TOP,padx=20,pady=15)
        e5.insert(0,'Mobile number')
        e5.bind('<FocusIn>',selfc.focusin)
        e6=Entry(f61,textvariable=selfc.signup_password,width=20,fg='grey',font=('times new roman',25,'bold'))
        e6.pack(side=TOP,padx=20,pady=15)
        e6.insert(0,'Password.')
        e6.bind('<FocusIn>',selfc.focusinpaswrd)
        e7=Entry(f61,textvariable=selfc.signup_cnfpassword,width=20,fg='grey',font=('times new roman',25,'bold'))
        e7.pack(side=TOP,padx=20,pady=15)
        e7.insert(0,'Confirm Password')
        e7.bind('<FocusIn>',selfc.focusinpaswrd)
        b2=Button(f4,text='Sign-up',font=('Calibri',20,'bold'),bg='green',fg='white',padx=10,pady=5,relief=RAISED,borderwidth=5)
        b2.place(relx=0.5,rely=0.8,anchor=CENTER)
        b2.bind('<Button-1>',selfc.signup_check)
        img=Image.open('money.gif')
        imgq=ImageTk.PhotoImage(img)
        l7=Label(f3,image=imgq)
        l7.place(relx=0.5,rely=0.5,anchor=CENTER)
        root.mainloop()

    def administrator(selfc):

        selfc.root.title('Admin')
        f0=Frame(selfc.root,bg='black')
        f0.pack(fill=BOTH,expand=True)
        f1=Frame(f0,bg='sky blue',padx=10)
        f1.pack(side=TOP,fill=X,padx=10,pady=5)
        f2=Frame(f0,bg='red',padx=5,pady=5)
        f2.pack(side=TOP,fill=BOTH,padx=2,pady=2,expand=1)
        f3=Frame(f2,bg='yellow',padx=5,pady=5)
        f3.pack(side=LEFT,fill=BOTH,expand=True,padx=2)
        f4=Frame(f2,bg='sky blue',padx=5,pady=5)
        f4.pack(side=LEFT,fill=BOTH,expand=True,padx=2)

        l1=Label(f1,text='Hey!! Welcome back Tanansh. What are you planning to do today?',
                 font=('Times new Roman',30,'bold'),bg='sky blue')
        l1.pack(side=TOP, fill=X)
        root.mainloop()

    def user(selfc):

        selfc.root.title(selfc.name)
        f0=Frame(selfc.root,bg='black')
        f0.pack(fill=BOTH,expand=True)
        f1=Frame(f0,bg='sky blue',padx=10)
        f1.pack(side=TOP,fill=X,padx=10,pady=5)
        f2=Frame(f0,bg='red',padx=5,pady=5)
        f2.pack(side=TOP,fill=BOTH,padx=2,pady=2,expand=1)
        f3=Frame(f2,bg='yellow',padx=5,pady=5)
        f3.pack(side=LEFT,fill=BOTH,expand=True,padx=2)
        f4=Frame(f2,bg='sky blue',padx=5,pady=5)
        f4.pack(side=LEFT,fill=BOTH,expand=True,padx=2)
#this line below this comment is important to see because i have used the '+' symbol which is used in java to concatenate two strings, but is here
#used in python, because if i try to run it by just using commas and string it was not happening, hence i needed to use this + symbol
        l1=Label(f1,text='Hey!! Welcome back '+selfc.name+'. What are you planning to do today?',
                 font=('Times new Roman',30,'bold'),bg='sky blue')
        l1.pack(side=TOP, fill=X)
        root.mainloop()

        

        
#main
root=Tk()
root.geometry('1350x700+0+0')
root.title('The Money Manager')
obj=creation(root)
obj.first_page()

