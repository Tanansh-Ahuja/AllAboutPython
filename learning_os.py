import os
import openpyxl

def save(j,k,lst):
    #print('\nWe are currently at ', os.getcwd())
    for i in lst:
        j =j +1
        if(os.path.isdir(i)):
            #print('\nWe are currently at ', os.getcwd())
            #k=k+1
            #print('\nRow and Column is: ',j,k)
            #print(i)
            sheet.cell(j,k).value = i
            wb.save("C:\\Users\\com\\Desktop\\movieslist.xlsx")
            os.chdir(i)
            q = list(os.listdir())
            save(j,k,q)
            #os.system("cd..")
            #k =k - 1
            #j=j-1
        else:
            if(i.endswith(".txt")):
                print('\n We have found a text file at ', os.getcwd())
                print('\nRow and Column is: ',j,k)
                sheet.cell(j,k).value = i
                wb.save("C:\\Users\\com\\Desktop\\movieslist.xlsx")
                #j =j +1





#main
wb = openpyxl.load_workbook("C:\\Users\\com\\Desktop\\movieslist.xlsx")
sheet = wb.active
print(os.getcwd())
os.chdir("folder")
print(os.getcwd())
#os.system("")
a = list(os.listdir())
save(0,3,a)
wb.save("C:\\Users\\com\\Desktop\\movieslist.xlsx")

        
    
    
