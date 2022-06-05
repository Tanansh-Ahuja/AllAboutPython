import openpyxl
w1=openpyxl.load_workbook('C:\\Users\\Tanansh Ahuja\\Desktop\\FINAL500.xlsx')
w2=openpyxl.load_workbook('C:\\Users\\Tanansh Ahuja\\Desktop\\sample2.xlsx')
sheet1=w1.active
sheet2=w2.active
#482

for i in range(6,482):
    a=sheet1.cell(i,1).value
    
    b=sheet1.cell(i,2).value
    if(b==None):
        continue
    try:
        h=sheet1.cell(i,2).hyperlink.target
    except:
        h='nothing'
    c=sheet1.cell(i,3).value
    sheet2.cell(i,1).value = a
    sheet2.cell(i,2).value = b
    sheet2.cell(i,3).hyperlink = h
    sheet2.cell(i,4).value=c
    
w2.save('C:\\Users\\Tanansh Ahuja\\Desktop\\sample2.xlsx')
